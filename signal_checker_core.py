from __future__ import annotations

import json
import re
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable, Optional

import pandas as pd


DEFAULT_OUTPUT_NAME = "signal_check_results.xlsx"

SIGNAL_LONG_COLUMN = "Signal long name"
SIGNAL_SHORT_COLUMN = "Signal short name"
SIGNAL_NAME_COLUMN = "Signal name"

# Internal canonical name used by the engine.
# The app now accepts either "Report event" or "Upload condition" in the reference file.
TRIGGER_COLUMN = "Report event"
REFERENCE_TRIGGER_ALIASES = ["Report event", "Upload condition"]

REFERENCE_REQUIRED_BASE = [SIGNAL_LONG_COLUMN, SIGNAL_SHORT_COLUMN]
LOGS_REQUIRED = ["triggerOrContext", "message"]

# Important: 0 is a valid value, not empty.
NULL_STRINGS = {"", "null", "none", "empty", "[null]", "nan", "n/a", "na"}


@dataclass
class CheckConfig:
    reference_path: str
    logs_path: str
    output_path: str
    export_debug: bool = True
    custom_path: Optional[str] = None


def read_table(file_path: str) -> pd.DataFrame:
    suffix = Path(file_path).suffix.lower()
    if suffix in {".xlsx", ".xls", ".xlsm"}:
        return pd.read_excel(file_path)
    if suffix == ".csv":
        return pd.read_csv(file_path)
    if suffix == ".tsv":
        return pd.read_csv(file_path, sep="\t")
    raise ValueError(f"Unsupported file type: {suffix}")


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\u00a0", " ").strip()
    text = re.sub(r"\s+", " ", text)
    return text


def canonical_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", normalize_text(value).lower())


def _normalized_columns(df: pd.DataFrame) -> list[str]:
    return [normalize_text(c) for c in df.columns]


def _find_column_name(df: pd.DataFrame, candidates: list[str]) -> Optional[str]:
    """
    Finds a column by exact normalized name first, then by canonical name.
    This allows small variations like extra spaces/case differences.
    """
    normalized_map = {normalize_text(c): c for c in df.columns}

    for candidate in candidates:
        candidate_norm = normalize_text(candidate)
        if candidate_norm in normalized_map:
            return normalized_map[candidate_norm]

    canonical_map = {canonical_key(c): c for c in df.columns}
    for candidate in candidates:
        candidate_key = canonical_key(candidate)
        if candidate_key in canonical_map:
            return canonical_map[candidate_key]

    return None


def _get_reference_trigger_source(df: pd.DataFrame) -> Optional[str]:
    return _find_column_name(df, REFERENCE_TRIGGER_ALIASES)


def normalize_reference_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """
    Normalizes reference files to the internal schema expected by the engine.

    Supported reference trigger columns:
      - Report event
      - Upload condition

    Internally both are mapped to:
      - Report event

    If both columns exist, "Report event" stays the primary source and empty
    values are filled from "Upload condition".
    """
    df = df.copy()
    df = df.rename(columns=lambda x: normalize_text(x))

    trigger_source = _get_reference_trigger_source(df)
    if trigger_source is None:
        return df

    trigger_source = normalize_text(trigger_source)

    if TRIGGER_COLUMN not in df.columns:
        df[TRIGGER_COLUMN] = df[trigger_source]
    elif trigger_source != TRIGGER_COLUMN and trigger_source in df.columns:
        # Keep Report event as the official engine column, but fill blanks from Upload condition.
        report_values = df[TRIGGER_COLUMN].fillna("").astype(str).map(normalize_text)
        fallback_values = df[trigger_source].fillna("").astype(str).map(normalize_text)
        df[TRIGGER_COLUMN] = report_values.where(report_values != "", fallback_values)

    # Keep a helper column for traceability in preview/debug if the reference used Upload condition.
    if trigger_source != TRIGGER_COLUMN:
        df["Reference trigger source"] = trigger_source

    return df


def is_null_like(value: Any) -> bool:
    if value is None:
        return True

    try:
        if pd.isna(value):
            return True
    except Exception:
        pass

    if isinstance(value, list):
        if len(value) == 0:
            return True
        return all(is_null_like(item) for item in value)

    if isinstance(value, tuple):
        if len(value) == 0:
            return True
        return all(is_null_like(item) for item in value)

    if isinstance(value, str):
        text = value.strip().lower()
        return text in NULL_STRINGS

    return False


def clean_for_excel(value: Any) -> str:
    if isinstance(value, (dict, list, tuple)):
        try:
            return json.dumps(value, ensure_ascii=False)
        except Exception:
            return str(value)
    if value is None:
        return ""
    return str(value)


def candidate_keys_from_reference(
    signal_long_name: str,
    signal_short_name: str = "",
    signal_name: str = "",
) -> list[tuple[str, str]]:
    """
    Returns deterministic candidates only.
    Each item: (candidate_key, strategy)
    """
    candidates: list[tuple[str, str]] = []
    seen: set[str] = set()

    def add(candidate: str, strategy: str) -> None:
        candidate = normalize_text(candidate)
        if candidate and candidate not in seen:
            seen.add(candidate)
            candidates.append((candidate, strategy))

    long_name = normalize_text(signal_long_name)
    short_name = normalize_text(signal_short_name)
    full_signal_name = normalize_text(signal_name)

    if long_name:
        add(long_name, "signal_long_name_exact")
        add(long_name[:1].lower() + long_name[1:] if long_name else long_name, "signal_long_name_lower_first")
        add(long_name.replace("_", ""), "signal_long_name_no_underscore")

        if long_name.startswith("DurationSpeedRange_"):
            suffix = long_name.replace("DurationSpeedRange_", "", 1)
            if suffix == "0":
                add("durationSpeedRange0", "duration_speed_special")
            elif suffix == "More_120":
                add("durationSpeedRangeMore120", "duration_speed_special")
            else:
                parts = suffix.split("_")
                if len(parts) == 2 and all(part.isdigit() for part in parts):
                    add(f"durationSpeedRangeBetween{parts[0]}And{parts[1]}", "duration_speed_special")

    if short_name:
        add(short_name, "signal_short_name_exact")

    if full_signal_name:
        add(full_signal_name.split("/")[-1], "signal_name_leaf")

    return candidates


def build_data_key_index(data_dict: dict[str, Any]) -> dict[str, str]:
    """canonical -> original"""
    index: dict[str, str] = {}
    for key in data_dict.keys():
        index[canonical_key(key)] = key
    return index


def find_signal_in_message(
    raw_message: str,
    signal_long_name: str,
    signal_short_name: str = "",
    signal_name: str = "",
) -> tuple[bool, Any, str, str]:
    """
    Returns:
        matched(bool),
        value(any),
        matched_key(str),
        strategy(str)
    """
    if not isinstance(raw_message, str) or not raw_message.strip():
        return False, None, "", "empty_message"

    try:
        payload = json.loads(raw_message)
    except Exception:
        return False, None, "", "invalid_json"

    data = payload.get("data")
    if not isinstance(data, dict):
        return False, None, "", "missing_data_dict"

    key_index = build_data_key_index(data)

    for candidate, strategy in candidate_keys_from_reference(signal_long_name, signal_short_name, signal_name):
        ckey = canonical_key(candidate)
        if ckey in key_index:
            real_key = key_index[ckey]
            return True, data.get(real_key), real_key, strategy

    return False, None, "", "no_matching_data_key"


def validate_reference_columns(df: pd.DataFrame) -> None:
    df = df.rename(columns=lambda x: normalize_text(x))

    missing_base = [
        required for required in REFERENCE_REQUIRED_BASE
        if _find_column_name(df, [required]) is None
    ]

    trigger_source = _get_reference_trigger_source(df)

    errors: list[str] = []
    if missing_base:
        errors.extend(missing_base)

    if trigger_source is None:
        errors.append("Report event OR Upload condition")

    if errors:
        raise ValueError(f"Reference file is missing required columns: {errors}")


def validate_logs_columns(df: pd.DataFrame) -> None:
    cols = {normalize_text(c) for c in df.columns}

    # Standard logs
    if "triggerOrContext" in cols and "message" in cols:
        return

    # DBB logs: we accept them if message exists, because triggerOrContext
    # can be extracted from the JSON payload.
    if "message" in cols:
        return

    missing = [c for c in LOGS_REQUIRED if normalize_text(c) not in cols]
    raise ValueError(f"Logs file is missing required columns: {missing}")


def _extract_trigger_from_message(raw_message: Any) -> str:
    """
    Extract trigger from DBB-style JSON message payload.
    Priority:
      1. context.triggerOrContext
      2. data.trigger
      3. data.standardTrigger
      4. root.triggerOrContext
    """
    if not isinstance(raw_message, str) or not raw_message.strip():
        return ""

    try:
        payload = json.loads(raw_message)
    except Exception:
        return ""

    context = payload.get("context", {})
    data = payload.get("data", {})

    trigger_value = ""

    if isinstance(context, dict):
        trigger_value = context.get("triggerOrContext", "") or ""

    if not trigger_value and isinstance(data, dict):
        trigger_value = data.get("trigger", "") or ""

    if not trigger_value and isinstance(data, dict):
        trigger_value = data.get("standardTrigger", "") or ""

    if not trigger_value:
        trigger_value = payload.get("triggerOrContext", "") or ""

    return normalize_text(trigger_value)


def enrich_dbb_logs_if_needed(df: pd.DataFrame) -> pd.DataFrame:
    """
    Supports DBB-style logs where triggerOrContext is embedded in JSON inside `message`.
    If `triggerOrContext` already exists, the dataframe is returned normalized only.
    """
    df = df.copy()
    df = df.rename(columns=lambda x: normalize_text(x))

    if "message" not in df.columns:
        return df

    if "triggerOrContext" in df.columns:
        df["triggerOrContext"] = df["triggerOrContext"].astype(str).map(normalize_text)
        return df

    extracted_triggers = [_extract_trigger_from_message(raw_message) for raw_message in df["message"]]
    df["triggerOrContext"] = extracted_triggers
    return df


def load_reference_file(path: str) -> pd.DataFrame:
    df = read_table(path)
    validate_reference_columns(df)
    df = normalize_reference_dataframe(df)

    # Final guard: downstream/UI code expects Report event.
    if TRIGGER_COLUMN not in df.columns:
        raise ValueError("Reference file could not be normalized. Required trigger column: Report event or Upload condition.")

    return df


def load_logs_file(path: str) -> pd.DataFrame:
    df = read_table(path)
    validate_logs_columns(df)

    df = df.rename(columns=lambda x: normalize_text(x))
    df = enrich_dbb_logs_if_needed(df)

    cols = {normalize_text(c) for c in df.columns}
    if "triggerOrContext" not in cols or "message" not in cols:
        raise ValueError(
            "Logs file could not be normalized. Required columns after normalization: "
            "triggerOrContext, message"
        )

    return df


def create_preview_text(reference_df: Optional[pd.DataFrame], logs_df: Optional[pd.DataFrame]) -> str:
    lines: list[str] = []

    if reference_df is not None:
        reference_df = normalize_reference_dataframe(reference_df)

        lines.append("=== REFERENCE (CCS) ===")
        lines.append(f"Rows: {len(reference_df)}")
        lines.append(f"Columns: {list(reference_df.columns)}")

        trigger_source = "Report event"
        if "Reference trigger source" in reference_df.columns:
            try:
                sources = sorted(
                    set(
                        str(v).strip()
                        for v in reference_df["Reference trigger source"].dropna().tolist()
                        if str(v).strip()
                    )
                )
                if sources:
                    trigger_source = " / ".join(sources)
            except Exception:
                trigger_source = "Report event"

        trigger_count = 0
        if TRIGGER_COLUMN in reference_df.columns:
            try:
                trigger_count = int((reference_df[TRIGGER_COLUMN].astype(str).str.strip() != "").sum())
            except Exception:
                trigger_count = 0

        lines.append(f"Reference trigger column used by engine: {TRIGGER_COLUMN}")
        lines.append(f"Original trigger source: {trigger_source}")
        lines.append(f"Rows with usable trigger: {trigger_count}")
        lines.append(reference_df.head(8).to_string(index=False))
        lines.append("")

    if logs_df is not None:
        lines.append("=== LOGS ===")
        lines.append(f"Rows: {len(logs_df)}")
        lines.append(f"Columns: {list(logs_df.columns)}")
        lines.append(logs_df.head(5).to_string(index=False))

        extracted = 0
        if "triggerOrContext" in logs_df.columns:
            try:
                extracted = int((logs_df["triggerOrContext"].astype(str).str.strip() != "").sum())
            except Exception:
                extracted = 0

        lines.append("")
        lines.append(f"Rows with extracted triggerOrContext: {extracted}")
        lines.append("")

    if reference_df is None and logs_df is None:
        lines.append("No files loaded yet.")

    return "\n".join(lines)




# ----------------------------------------------------------------------
# Custom applicability support
# ----------------------------------------------------------------------
def _split_custom_name(raw_name: Any) -> dict[str, str]:
    """
    Custom applicability entries are encoded as:
      shortName;trigger;customSetting;uid;freshness;privacy

    The first field is always the signal short name. The trigger and the
    remaining fields are kept for traceability, but the applicability key is
    the short name.
    """
    raw = normalize_text(raw_name)
    parts = [normalize_text(part) for part in raw.split(";")]
    while len(parts) < 6:
        parts.append("")

    return {
        "Custom Raw Name": raw,
        "Custom Signal short name": parts[0],
        "Custom Trigger": parts[1],
        "Custom Setting": parts[2],
        "Custom UID": parts[3],
        "Custom Freshness": parts[4],
        "Custom Privacy": parts[5],
    }


def _load_custom_xml(path: str) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    tree = ET.parse(path)
    root = tree.getroot()

    custom_setting_name = normalize_text(root.attrib.get("Name", ""))
    ecu = normalize_text(root.attrib.get("ECU", ""))
    action_type = normalize_text(root.attrib.get("ActionType", ""))
    method = normalize_text(root.attrib.get("Method", ""))
    status = normalize_text(root.attrib.get("status", ""))

    for prop in root.findall(".//property"):
        parsed = _split_custom_name(prop.attrib.get("Name", ""))
        if not parsed["Custom Signal short name"]:
            continue

        parsed.update(
            {
                "Custom Value": normalize_text(prop.attrib.get("value", "")),
                "Custom Type": normalize_text(prop.attrib.get("Type", "")),
                "Custom Root Setting": custom_setting_name,
                "Custom ECU": ecu,
                "Custom ActionType": action_type,
                "Custom Method": method,
                "Custom Root Status": status,
                "Custom Source Type": "XML",
                "Custom Source File": Path(path).name,
            }
        )
        rows.append(parsed)

    return pd.DataFrame(rows)


def _load_custom_csv(path: str) -> pd.DataFrame:
    df = read_table(path).rename(columns=lambda x: normalize_text(x))
    rows: list[dict[str, Any]] = []

    # Preferred format: a Name column containing the semicolon-encoded value.
    name_col = _find_column_name(df, ["Name", "property Name", "Custom Name", "Property Name"])

    # Alternative format: direct columns already split in the CSV.
    short_col = _find_column_name(df, ["Signal short name", "Short Name", "ShortName", "Signal Short Name", "Custom Signal short name"])
    trigger_col = _find_column_name(df, ["Trigger", "Report event", "Upload condition", "Custom Trigger"])
    setting_col = _find_column_name(df, ["Custom Setting", "Setting", "CustomSetting", "Custom Root Setting"])
    uid_col = _find_column_name(df, ["UID", "Custom UID"])
    freshness_col = _find_column_name(df, ["Freshness", "Custom Freshness"])
    privacy_col = _find_column_name(df, ["Privacy", "Custom Privacy"])
    value_col = _find_column_name(df, ["value", "Value", "Custom Value"])

    # If no explicit Name column exists, search for the first column with semicolon-encoded values.
    if name_col is None:
        for col in df.columns:
            sample = df[col].dropna().astype(str).head(20).tolist()
            if any(";" in item for item in sample):
                name_col = col
                break

    for _, row in df.fillna("").iterrows():
        if name_col is not None and normalize_text(row.get(name_col, "")):
            parsed = _split_custom_name(row.get(name_col, ""))
        else:
            short_name = normalize_text(row.get(short_col, "")) if short_col else ""
            if not short_name:
                continue
            raw_name = ";".join(
                [
                    short_name,
                    normalize_text(row.get(trigger_col, "")) if trigger_col else "",
                    normalize_text(row.get(setting_col, "")) if setting_col else "",
                    normalize_text(row.get(uid_col, "")) if uid_col else "",
                    normalize_text(row.get(freshness_col, "")) if freshness_col else "",
                    normalize_text(row.get(privacy_col, "")) if privacy_col else "",
                ]
            )
            parsed = _split_custom_name(raw_name)

        if not parsed["Custom Signal short name"]:
            continue

        parsed.update(
            {
                "Custom Value": normalize_text(row.get(value_col, "")) if value_col else "",
                "Custom Type": "",
                "Custom Root Setting": normalize_text(row.get(setting_col, "")) if setting_col else parsed.get("Custom Setting", ""),
                "Custom ECU": "",
                "Custom ActionType": "",
                "Custom Method": "",
                "Custom Root Status": "",
                "Custom Source Type": "CSV",
                "Custom Source File": Path(path).name,
            }
        )
        rows.append(parsed)

    return pd.DataFrame(rows)


def load_custom_file(path: str) -> pd.DataFrame:
    suffix = Path(path).suffix.lower()
    if suffix == ".xml":
        return _load_custom_xml(path)
    if suffix in {".csv", ".tsv", ".xlsx", ".xls", ".xlsm"}:
        return _load_custom_csv(path)
    raise ValueError(f"Unsupported custom applicability file type: {suffix}")


def _best_summary_row(
    summary_rows: list[dict[str, Any]],
    custom_trigger: str = "",
) -> dict[str, Any] | None:
    """
    Choose the best Summary row for one Custom applicability entry.

    Custom is the applicability reference, so the Custom row is never rejected
    because of trigger. However, when the same short name exists in Summary with
    the same trigger as the Custom row, that Summary row is preferred for clearer
    traceability. If no same-trigger Summary row exists, the function falls back
    to the best row by status/occurrences.
    """
    if not summary_rows:
        return None

    custom_trigger_key = canonical_key(custom_trigger)

    def priority(row: dict[str, Any]) -> tuple[int, int, int, int]:
        status = normalize_text(row.get("Status", ""))
        signal_occ = int(row.get("Signal Occurrences", 0) or 0)
        trigger_occ = int(row.get("Trigger Occurrences", 0) or 0)
        summary_trigger_key = canonical_key(row.get("Trigger", ""))

        # Prefer the same trigger when available, but do not require it.
        trigger_match_penalty = 0 if custom_trigger_key and summary_trigger_key == custom_trigger_key else 1

        if status == "FOUND":
            status_rank = 0
        elif status == "SIGNAL NOT FOUND":
            status_rank = 1
        elif status == "TRIGGER NOT FOUND":
            status_rank = 2
        else:
            status_rank = 3

        return (trigger_match_penalty, status_rank, -signal_occ, -trigger_occ)

    return sorted(summary_rows, key=priority)[0]


def build_custom_applicability_sheet(
    custom_df: pd.DataFrame,
    summary_df: pd.DataFrame,
) -> pd.DataFrame:
    """
    Build the Custom Applicability sheet using Custom XML/CSV as the reference.

    Agreed logic:
      - Custom.xml / Custom.csv = applicability reference for the exact car.
      - CCS_Parameters = main base/reference where signal metadata is found.
      - Logs.csv / 323.csv = what was performed on vehicle.

    For every Custom entry, use the first field before ';' as Signal Short Name.
    Match that short name to Summary[Signal short name]. The Custom trigger is
    shown and used only to choose the clearest Summary row when available; it is
    not used to reject applicability.
    """
    if custom_df is None or custom_df.empty:
        return pd.DataFrame()

    summary = summary_df.copy().fillna("")
    custom = custom_df.copy().fillna("")

    if "Signal short name" not in summary.columns:
        summary["Signal short name"] = ""

    summary["_short_key"] = summary["Signal short name"].astype(str).map(canonical_key)
    custom["_short_key"] = custom["Custom Signal short name"].astype(str).map(canonical_key)

    summary_by_short: dict[str, list[dict[str, Any]]] = {}
    for _, row in summary.iterrows():
        key = row.get("_short_key", "")
        if key:
            summary_by_short.setdefault(key, []).append(row.to_dict())

    out_rows: list[dict[str, Any]] = []

    for _, custom_row in custom.iterrows():
        short_key = custom_row.get("_short_key", "")
        custom_trigger = normalize_text(custom_row.get("Custom Trigger", ""))
        summary_rows = summary_by_short.get(short_key, [])
        best = _best_summary_row(summary_rows, custom_trigger=custom_trigger)

        if best is None:
            summary_status = ""
            result_status = "CUSTOM SIGNAL APPLICABLE BUT NOT FOUND IN CCS SUMMARY"
            best = {}
        else:
            summary_status = normalize_text(best.get("Status", ""))

            if summary_status == "FOUND":
                result_status = "PRESENT IN SUMMARY AND APPLICABLE IN CUSTOM.XML"
            elif summary_status == "SIGNAL NOT FOUND":
                result_status = "SIGNAL NOT PRESENT IN SUMMARY AND APPLICABLE IN CUSTOM.XML"
            elif summary_status == "TRIGGER NOT FOUND":
                result_status = "TRIGGER NOT PRESENT IN SUMMARY BUT APPLICABLE IN CUSTOM.XML"
            else:
                result_status = "CUSTOM SIGNAL APPLICABLE BUT SUMMARY STATUS UNKNOWN"

        custom_short = custom_row.get("Custom Signal short name", "")
        summary_trigger = best.get("Trigger", "")
        summary_signal_short = best.get("Signal short name", custom_short)

        out_rows.append(
            {
                # Put the important business status early so it is easy to see in Excel.
                "Status": result_status,
                "Custom Applicability Status": result_status,
                "Custom Signal short name": custom_short,
                "Custom Trigger": custom_trigger,
                "Custom Setting": custom_row.get("Custom Setting", ""),
                "Custom UID": custom_row.get("Custom UID", ""),
                "Custom Freshness": custom_row.get("Custom Freshness", ""),
                "Custom Privacy": custom_row.get("Custom Privacy", ""),
                "Custom Value": custom_row.get("Custom Value", ""),
                "Custom Source Type": custom_row.get("Custom Source Type", ""),
                "Custom Source File": custom_row.get("Custom Source File", ""),
                "Custom Raw Name": custom_row.get("Custom Raw Name", ""),
                "Summary Trigger": summary_trigger,
                "Signal long name": best.get("Signal long name", ""),
                "Signal short name": summary_signal_short,
                "Signal name": best.get("Signal name", ""),
                "Trigger Occurrences": best.get("Trigger Occurrences", ""),
                "Signal Occurrences": best.get("Signal Occurrences", ""),
                "Null/Empty Occurrences": best.get("Null/Empty Occurrences", ""),
                "Present in Logs": best.get("Present in Logs", "NO"),
                "Matched Key(s)": best.get("Matched Key(s)", ""),
                "Match Strategy": best.get("Match Strategy", ""),
                "Values Seen": best.get("Values Seen", ""),
                "Summary Status": summary_status,
                "Custom Logic": "Custom file is the car applicability reference. Every Custom row is exported. Matching to CCS/Summary is by Signal short name. Custom trigger is kept for traceability and only used to prefer the same Summary trigger when available; it never rejects applicability.",
            }
        )

    return pd.DataFrame(out_rows)


def run_signal_check(
    reference_df: pd.DataFrame,
    logs_df: pd.DataFrame,
    config: CheckConfig,
    progress_callback: Optional[Callable[[str], None]] = None,
    log_callback: Optional[Callable[[str], None]] = None,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    reference_df = normalize_reference_dataframe(reference_df.copy())
    logs_df = logs_df.copy()

    reference_df = reference_df.fillna("")
    logs_df = logs_df.fillna("")

    if TRIGGER_COLUMN not in reference_df.columns:
        raise ValueError("Reference file must contain either 'Report event' or 'Upload condition'.")

    reference_df = reference_df[
        (reference_df[TRIGGER_COLUMN].astype(str).str.strip() != "")
        & (reference_df[SIGNAL_LONG_COLUMN].astype(str).str.strip() != "")
    ].copy()

    logs_df["triggerOrContext"] = logs_df["triggerOrContext"].astype(str).map(normalize_text)
    logs_df["message"] = logs_df["message"].astype(str)

    summary_rows: list[dict[str, Any]] = []
    debug_rows: list[dict[str, Any]] = []

    total = len(reference_df)

    if log_callback:
        log_callback("Initializing analysis engine...")
        log_callback(f"Reference rows eligible for processing: {total}")
        log_callback(f"Log rows loaded: {len(logs_df)}")
        extracted = int((logs_df["triggerOrContext"].astype(str).str.strip() != "").sum())
        log_callback(f"Rows with usable triggerOrContext: {extracted}")

    for idx, (_, ref_row) in enumerate(reference_df.iterrows(), start=1):
        trigger = normalize_text(ref_row[TRIGGER_COLUMN])
        signal_long = normalize_text(ref_row[SIGNAL_LONG_COLUMN])
        signal_short = normalize_text(ref_row.get(SIGNAL_SHORT_COLUMN, ""))
        signal_name = normalize_text(ref_row.get(SIGNAL_NAME_COLUMN, ""))

        if progress_callback:
            progress_callback(f"Processing {idx}/{total}: {signal_long}")

        trigger_logs = logs_df[logs_df["triggerOrContext"] == trigger].copy()
        trigger_occurrences = len(trigger_logs)

        signal_occurrences = 0
        null_occurrences = 0
        values_seen: list[str] = []
        matched_keys: set[str] = set()
        strategies: set[str] = set()

        for occ_idx, (_, log_row) in enumerate(trigger_logs.iterrows(), start=1):
            matched, value, matched_key, strategy = find_signal_in_message(
                raw_message=log_row["message"],
                signal_long_name=signal_long,
                signal_short_name=signal_short,
                signal_name=signal_name,
            )

            is_null = False
            if matched:
                signal_occurrences += 1
                matched_keys.add(matched_key)
                strategies.add(strategy)

                if is_null_like(value):
                    null_occurrences += 1
                    is_null = True

                values_seen.append(clean_for_excel(value))

            if config.export_debug:
                debug_rows.append(
                    {
                        "Trigger": trigger,
                        "Signal long name": signal_long,
                        "Signal short name": signal_short,
                        "Signal name": signal_name,
                        "Occurrence Index": occ_idx,
                        "Trigger Occurrences": trigger_occurrences,
                        "Matched Signal": "YES" if matched else "NO",
                        "Matched Key": matched_key,
                        "Match Strategy": strategy,
                        "Extracted Value": clean_for_excel(value),
                        "Is Null / Empty": "YES" if is_null else "NO",
                        "Raw Message": log_row["message"],
                    }
                )

        present = "YES" if signal_occurrences > 0 else "NO"

        if trigger_occurrences == 0:
            status = "TRIGGER NOT FOUND"
        elif signal_occurrences == 0:
            status = "SIGNAL NOT FOUND"
        else:
            status = "FOUND"

        summary_rows.append(
            {
                "Trigger": trigger,
                "Signal long name": signal_long,
                "Signal short name": signal_short,
                "Signal name": signal_name,
                "Trigger Occurrences": trigger_occurrences,
                "Signal Occurrences": signal_occurrences,
                "Null/Empty Occurrences": null_occurrences,
                "Present in Logs": present,
                "Matched Key(s)": " | ".join(sorted(k for k in matched_keys if k)),
                "Match Strategy": " | ".join(sorted(s for s in strategies if s)),
                "Values Seen": " | ".join(values_seen),
                "Status": status,
            }
        )

    summary_df = pd.DataFrame(summary_rows)
    debug_df = pd.DataFrame(debug_rows)

    custom_applicability_df: Optional[pd.DataFrame] = None
    custom_path = normalize_text(getattr(config, "custom_path", ""))
    if custom_path:
        if log_callback:
            log_callback(f"Loading Custom applicability reference: {custom_path}")
        custom_df = load_custom_file(custom_path)
        custom_applicability_df = build_custom_applicability_sheet(custom_df, summary_df)
        if log_callback:
            log_callback(f"Custom applicability rows: {len(custom_applicability_df)}")

    write_output(Path(config.output_path), summary_df, debug_df, config.export_debug, custom_applicability_df)

    if log_callback:
        log_callback("Analysis completed.")
        log_callback(f"Summary rows: {len(summary_df)}")
        if config.export_debug:
            log_callback(f"Debug rows: {len(debug_df)}")
        log_callback(f"Saved to: {config.output_path}")

    return summary_df, debug_df


def write_output(
    output_path: Path,
    summary_df: pd.DataFrame,
    debug_df: pd.DataFrame,
    export_debug: bool,
    custom_applicability_df: Optional[pd.DataFrame] = None,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, index=False, sheet_name="Summary")

        if export_debug:
            debug_df.to_excel(writer, index=False, sheet_name="Debug")

        if custom_applicability_df is not None and not custom_applicability_df.empty:
            custom_applicability_df.to_excel(writer, index=False, sheet_name="Custom Applicability")

        readme = pd.DataFrame(
            [
                {
                    "Field": "Logic",
                    "Value": "Rows are matched by reference trigger column = logs triggerOrContext. The reference trigger can be Report event or Upload condition.",
                },
                {
                    "Field": "Reference trigger aliases",
                    "Value": "Accepted reference columns: Report event, Upload condition. Internally both are normalized to Report event.",
                },
                {
                    "Field": "Signal match",
                    "Value": "Signal is searched only inside JSON data keys, using deterministic aliases from Signal long name / Signal name / Signal short name.",
                },
                {"Field": "Occurrence", "Value": "Signal Occurrences = number of trigger messages where the signal key was found."},
                {"Field": "Values", "Value": "Values Seen contains all extracted values in occurrence order."},
                {"Field": "Null rule", "Value": "0 is valid. None, empty string, EMPTY, NaN, [null], [] are treated as null/empty."},
                {"Field": "Status", "Value": "FOUND / SIGNAL NOT FOUND / TRIGGER NOT FOUND."},
                {
                    "Field": "DBB support",
                    "Value": "If triggerOrContext column is missing, it is extracted automatically from JSON message payload (context.triggerOrContext, data.trigger, data.standardTrigger).",
                },
                {
                    "Field": "Custom applicability",
                    "Value": "Custom XML/CSV is treated as the applicability reference for the exact car. Every Custom row is exported. The first field in Custom Name is Signal Short Name. Matching to Summary/CCS uses Signal short name. Custom trigger is displayed and used only to prefer the same Summary trigger when available; it never rejects applicability.",
                },
            ]
        )
        readme.to_excel(writer, index=False, sheet_name="Read Me")

    auto_format_excel(output_path)


def auto_format_excel(output_path: Path) -> None:
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill

        wb = load_workbook(output_path)

        header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
        found_fill = PatternFill(fill_type="solid", fgColor="E8F5E9")
        missing_fill = PatternFill(fill_type="solid", fgColor="FDE9E7")
        warn_fill = PatternFill(fill_type="solid", fgColor="FFF4CC")

        for ws in wb.worksheets:
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = header_fill

            for column_cells in ws.columns:
                max_len = 0
                letter = column_cells[0].column_letter

                for cell in column_cells:
                    value = "" if cell.value is None else str(cell.value)
                    if len(value) > max_len:
                        max_len = len(value)

                ws.column_dimensions[letter].width = min(max_len + 2, 60)

            if ws.title == "Summary":
                headers = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}
                status_col = headers.get("Status")

                if status_col:
                    for row in range(2, ws.max_row + 1):
                        cell = ws.cell(row=row, column=status_col)
                        if cell.value == "FOUND":
                            cell.fill = found_fill
                        elif cell.value == "SIGNAL NOT FOUND":
                            cell.fill = warn_fill
                        elif cell.value == "TRIGGER NOT FOUND":
                            cell.fill = missing_fill

            if ws.title == "Custom Applicability":
                headers = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}
                status_col = headers.get("Status") or headers.get("Custom Applicability Status")

                if status_col:
                    for row in range(2, ws.max_row + 1):
                        cell = ws.cell(row=row, column=status_col)
                        text_value = "" if cell.value is None else str(cell.value)
                        if "PRESENT IN SUMMARY" in text_value:
                            cell.fill = found_fill
                        elif "SIGNAL NOT PRESENT" in text_value or "TRIGGER NOT PRESENT" in text_value:
                            cell.fill = warn_fill
                        else:
                            cell.fill = missing_fill

        wb.save(output_path)
    except Exception:
        pass
